import  openpyxl as xl
from openpyxl.chart import BarChart,Reference

def analyse(filename):
    workbook = xl.load_workbook(filename)
    sheet1 = workbook["Sheet1"]
    cell = sheet1.cell(1, 1)

    for row in range(2, sheet1.max_row + 1):
        cell1 = sheet1.cell(row, 2)

        deviation = 500 - cell1.value
        print(deviation)
        deviation_cell = sheet1.cell(row, 4)
        deviation_cell.value = deviation

    values = Reference(
        sheet1,
        min_row=2,
        max_row=sheet1.max_row,
        min_col=4,
        max_col=4
    )

    chart1 = BarChart()
    chart1.add_data(values)
    sheet1.add_chart(chart1, "e5")

    workbook.save(filename)

analyse("students.xlsx")